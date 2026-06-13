from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

class ExcelExporter:
    @staticmethod
    def export_ledger(records, file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "台账流水"
        
        headers = ['时间', '类型', '商品名称', '数量', '备注', '操作人']
        
        header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record['created_at']).border = thin_border
            ws.cell(row=row, column=2, value=record['type']).border = thin_border
            ws.cell(row=row, column=3, value=record['product_name']).border = thin_border
            qty = f"+{record['quantity']}" if record['type'] == '入库' else f"-{record['quantity']}"
            ws.cell(row=row, column=4, value=qty).border = thin_border
            ws.cell(row=row, column=5, value=record.get('remark', '')).border = thin_border
            ws.cell(row=row, column=6, value=record['operator_name']).border = thin_border
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        ws.freeze_panes = "A2"
        
        directory = os.path.dirname(file_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory)
        
        wb.save(file_path)
        return True
